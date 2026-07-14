"""Launch spec 2026-07-14 — sections 2c + 3 backend.

What this file proves:
  - OWNER GAP-ONLY BOARD: under the owner view (X-View: owner or ?view=owner)
    /api/top100 and /api/top100/funnel return ONLY gap stores
    (skus_carried <= 1), filtered BEFORE the 100 cap. Internal view is
    unchanged; internal ?gap_only=1 gives parity. The gap board and the full
    board never share a cache entry.
  - EXCEL ROUND-TRIP: the top100.xlsx export carries the round-trip columns
    (Store # key, Rank editable, Owner Status, SKUs Carried) and
    POST /api/top100/import-xlsx applies changed ranks/statuses through the
    SAME audited resequence path (changed_by 'owner-xlsx'/'internal-xlsx'),
    skips unknown stores and bad values with reasons, keeps ranks unique
    sequential, invalidates the cache, and is owner-allowlisted.
  - CONTACTS HISTORY (mistake-proofing): PUT /api/stores/<id> audits every
    contact-field change old -> new into territory_status_history, and
    GET /api/crm/store/<n>/contacts/history returns per-field current value
    + up to 10 prior changes + the restore candidate. Restore is a normal
    audited update; blanking a name keeps the prior value one tap away. The
    endpoint is INTERNAL ONLY — the owner view gets 403 (fail-closed).

Fixture: the same 6 hand-built stores as test_top100_v2.py, so the carried
facts are known exactly:

  store  carries              skus_carried   gap?
  101    nothing              0              yes
  102    nothing              0              yes
  103    Phoenix (SOD 5)      1              yes
  104    BOTH (SOD+live)      2              NO — the only non-gap store
  105    live 0 (fresher)     0              yes
  106    Phoenix (live 3)     1              yes

Run with: python3 -m pytest tests/test_launch_v3.py -v
"""
import io
import os
import sqlite3
import sys

import pytest

# Force SQLite for tests so we never touch production Postgres.
os.environ.pop('DATABASE_URL', None)
os.environ.pop('ADMIN_TOKEN', None)
os.environ.pop('API_KEY', None)
os.environ.pop('SOD_CRON_TOKEN', None)

TEST_DB_DIR = '/tmp/drippcan_launch_v3_test'
os.makedirs(TEST_DB_DIR, exist_ok=True)
TEST_DB = os.path.join(TEST_DB_DIR, 'drippcan.db')

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

PHOENIX = '0014318'
DAYAA = '0044451'
APP_PY = os.path.join(os.path.dirname(__file__), '..', 'app.py')

SOD_DATE = '2026-07-13'            # start-of-day snapshot
LIVE_AT = '2026-07-14 09:00:00'    # fresher live batch (next morning)

GAP_STORES = {101, 102, 103, 105, 106}   # skus_carried <= 1
NON_GAP_STORE = 104                       # carries BOTH SKUs

#         sn,  city,          class, tier,         route_day, route_stop
FIXTURE_STORES = [
    (101, 'Toronto',     'B',   'routed',     1, 1),
    (102, 'Mississauga', 'AAA', 'routed',     1, 2),
    (103, 'Toronto',     'AAA', 'routed',     2, 1),
    (104, 'Brampton',    'AA',  'territory',  None, None),
    (105, 'Vaughan',     'A',   'territory',  None, None),
    (106, 'Scarborough', '',    'discovered', None, None),
]

# CRM stores row for the contacts-history tests. Startup seeds the full
# 766-store book into `stores`, so store 101 already exists — the fixture
# resolves its id instead of inserting a duplicate.
CRM_STORE_SN = 101


@pytest.fixture(scope='module')
def app_module():
    """Import app.py fresh against an isolated SQLite file (same pattern as
    test_top100_v2.py — DB_DIR re-asserted because collection order matters)."""
    prev_db_dir = os.environ.get('DB_DIR')
    os.environ['DB_DIR'] = TEST_DB_DIR
    if os.path.exists(TEST_DB):
        os.remove(TEST_DB)
    for mod in list(sys.modules):
        if mod == 'app' or mod.startswith('app.'):
            del sys.modules[mod]
    import importlib.util
    spec = importlib.util.spec_from_file_location('app', APP_PY)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    yield m
    if prev_db_dir is None:
        os.environ.pop('DB_DIR', None)
    else:
        os.environ['DB_DIR'] = prev_db_dir


@pytest.fixture
def client(app_module):
    app_module.app.config['TESTING'] = True
    app_module._rate_buckets.clear()
    with app_module.app.test_client() as c:
        yield c


def _db():
    conn = sqlite3.connect(TEST_DB)
    conn.row_factory = sqlite3.Row
    return conn


def _rank_state():
    db = _db()
    out = {r['store_number']: r['priority_rank'] for r in db.execute(
        "SELECT store_number, priority_rank FROM territory_stores "
        "WHERE active=1").fetchall()}
    db.close()
    return out


def _assert_unique_sequential(state):
    ranks = sorted(v for v in state.values() if v is not None)
    assert ranks == list(range(1, len(ranks) + 1)), (
        f'ranks must be unique sequential 1..N, got {ranks}')


def _history_rows(field, store_number=None):
    db = _db()
    q = ("SELECT store_number, field, old_value, new_value, changed_by "
         "FROM territory_status_history WHERE field=?")
    params = [field]
    if store_number is not None:
        q += " AND store_number=?"
        params.append(store_number)
    q += " ORDER BY id"
    rows = [dict(r) for r in db.execute(q, params).fetchall()]
    db.close()
    return rows


@pytest.fixture(scope='module')
def fixture_stores(app_module):
    """The 6 designed territory stores + known SOD/live rows + one CRM
    stores row (contacts tests need PUT /api/stores/<id> to resolve)."""
    db = _db()
    for sn, city, klass, tier, rd, rs in FIXTURE_STORES:
        db.execute(
            "INSERT INTO territory_stores "
            "(store_number, tier, class, account, address, city, postal, "
            " route_day, route_stop, source) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (sn, tier, klass, f'LCBO #{sn}', f'{sn} Test Rd', city, '',
             rd, rs, 'seed'))
    sod_rows = [
        (PHOENIX, 103, 5, 'PHOENIX ULTRA SMOOTH VODKA'),   # 103 carries one
        (PHOENIX, 104, 6, 'PHOENIX ULTRA SMOOTH VODKA'),   # 104 carries both
        (DAYAA,   104, 4, 'DAYAA ARAK'),
        (PHOENIX, 105, 4, 'PHOENIX ULTRA SMOOTH VODKA'),   # superseded by live 0
    ]
    for sku, sn, on_hand, name in sod_rows:
        db.execute(
            "INSERT INTO sod_inventory "
            "(sku, store_number, snapshot_date, status, on_hand, product_name) "
            "VALUES (?,?,?,?,?,?)", (sku, sn, SOD_DATE, 'L', on_hand, name))
    live_rows = [
        (PHOENIX, 105, 0),   # fresher live says SOLD OUT -> 105 is a gap store
        (PHOENIX, 106, 3),   # live-only listing -> 106 carries one
        (DAYAA,   104, 2),   # fresher reading for 104's Dayaa
    ]
    for sku, sn, qty in live_rows:
        db.execute(
            "INSERT INTO lcbo_live_snapshots "
            "(sku, store_number, qty, store_name, city, checked_at, batch_id) "
            "VALUES (?,?,?,?,?,?,?)",
            (sku, sn, qty, f'LCBO #{sn}', '', LIVE_AT, 'batch-launch-v3'))
    db.commit()
    db.close()
    return [s[0] for s in FIXTURE_STORES]


@pytest.fixture(scope='module')
def crm_store_id(app_module):
    """id of the seeded stores row for CRM_STORE_SN (PUT /api/stores/<id>)."""
    db = _db()
    row = db.execute("SELECT id FROM stores WHERE store_number=?",
                     (CRM_STORE_SN,)).fetchone()
    db.close()
    assert row is not None, f'store {CRM_STORE_SN} missing from seeded book'
    return row['id']


def _board(client, headers=None, qs=''):
    resp = client.get(f'/api/top100{qs}', headers=headers or {})
    assert resp.status_code == 200, resp.get_data(as_text=True)
    return resp.get_json()


def _board_stores(body):
    return {int(r['store_number']) for r in body['rows']}


class TestOwnerGapOnlyBoard:
    """Section 3: owner top-100 = gap stores only, internal unchanged."""

    def test_internal_board_is_full(self, fixture_stores, client):
        body = _board(client)
        assert body['gap_only'] is False
        assert body['count'] == 6
        assert _board_stores(body) == GAP_STORES | {NON_GAP_STORE}

    def test_owner_board_is_gap_only_via_header(self, fixture_stores, client):
        body = _board(client, headers={'X-View': 'owner'})
        assert body['gap_only'] is True
        assert body['count'] == 5
        assert _board_stores(body) == GAP_STORES
        assert all(r['skus_carried'] <= 1 for r in body['rows'])

    def test_owner_board_is_gap_only_via_query(self, fixture_stores, client):
        body = _board(client, qs='?view=owner')
        assert body['gap_only'] is True
        assert _board_stores(body) == GAP_STORES

    def test_internal_gap_only_param_parity(self, fixture_stores, client):
        body = _board(client, qs='?gap_only=1')
        assert body['gap_only'] is True
        assert _board_stores(body) == GAP_STORES
        # Internal gap view keeps internal fields intact (no owner scrub).
        assert body['count'] == 5

    def test_gap_and_full_boards_never_share_cache(self, fixture_stores, client):
        # Warm the FULL board cache, then hit the owner view: it must get the
        # filtered board, not the cached full body. Then internal again must
        # still be full — key_extra keeps the two on separate cache keys.
        full = _board(client)
        assert _board_stores(full) == GAP_STORES | {NON_GAP_STORE}
        owner = _board(client, headers={'X-View': 'owner'})
        assert _board_stores(owner) == GAP_STORES
        full_again = _board(client)
        assert _board_stores(full_again) == GAP_STORES | {NON_GAP_STORE}

    def test_funnel_owner_counts_gap_stores_only(self, fixture_stores, client):
        # Move the NON-gap store through the funnel internally...
        resp = client.post('/api/top100/status', json={
            'store_number': NON_GAP_STORE, 'owner_status': 'order_received',
            'changed_by': 'Namit'})
        assert resp.status_code == 200, resp.get_data(as_text=True)
        # ...internal funnel sees all 6 including the order...
        internal = client.get('/api/top100/funnel').get_json()
        assert internal['gap_only'] is False
        assert internal['board_size'] == 6
        assert internal['funnel']['order_received'] == 1
        # ...but the owner funnel counts ONLY the 5 gap stores.
        owner = client.get('/api/top100/funnel',
                           headers={'X-View': 'owner'}).get_json()
        assert owner['gap_only'] is True
        assert owner['board_size'] == 5
        assert owner['funnel']['order_received'] == 0
        assert owner['funnel']['none'] == 5


def _download_xlsx(client, headers=None):
    resp = client.get('/api/export/top100.xlsx', headers=headers or {})
    assert resp.status_code == 200, resp.get_data(as_text=True)
    return resp.data


def _sheet_rows(blob):
    from openpyxl import load_workbook
    ws = load_workbook(io.BytesIO(blob)).active
    rows = list(ws.iter_rows(values_only=True))
    return list(rows[0]), rows[1:]


def _post_xlsx(client, blob, headers=None):
    return client.post(
        '/api/top100/import-xlsx',
        data={'file': (io.BytesIO(blob), 'top100.xlsx')},
        content_type='multipart/form-data',
        headers=headers or {})


def _build_xlsx(header, rows):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestXlsxRoundTrip:
    """Section 3: top100.xlsx export columns + audited import round-trip."""

    def test_export_has_round_trip_columns(self, fixture_stores, client):
        header, rows = _sheet_rows(_download_xlsx(client))
        for col in ('Rank', 'Store #', 'Account', 'City',
                    'Owner Status', 'SKUs Carried'):
            assert col in header, f'missing round-trip column {col!r}'
        by_store = {r[header.index('Store #')]: r for r in rows}
        carried_idx = header.index('SKUs Carried')
        assert by_store[NON_GAP_STORE][carried_idx] == 2
        assert by_store[101][carried_idx] == 0
        assert by_store[103][carried_idx] == 1

    def test_import_unmodified_sheet_is_a_noop(self, fixture_stores, client):
        resp = _post_xlsx(client, _download_xlsx(client))
        assert resp.status_code == 200, resp.get_data(as_text=True)
        body = resp.get_json()
        assert body['updated'] == 0
        assert body['skipped'] == []

    def test_import_applies_rank_change_audited(self, fixture_stores, client):
        blob = _download_xlsx(client)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(blob))
        ws = wb.active
        header = [c.value for c in ws[1]]
        rank_col = header.index('Rank') + 1
        store_col = header.index('Store #') + 1
        target_row = next(r for r in range(2, ws.max_row + 1)
                          if ws.cell(row=r, column=store_col).value == 105)
        ws.cell(row=target_row, column=rank_col).value = 1
        buf = io.BytesIO()
        wb.save(buf)

        resp = _post_xlsx(client, buf.getvalue())
        assert resp.status_code == 200, resp.get_data(as_text=True)
        body = resp.get_json()
        assert body['updated'] == 1
        assert body['skipped'] == []
        assert body['changed_by'] == 'internal-xlsx'
        assert body['total_ranked'] == 1

        state = _rank_state()
        assert state[105] == 1
        _assert_unique_sequential(state)
        audits = _history_rows('priority_rank', 105)
        assert audits[-1]['changed_by'] == 'internal-xlsx'
        assert audits[-1]['new_value'] == '1'
        # Cache invalidated: the fresh board shows the new rank immediately.
        board = _board(client)
        row_105 = next(r for r in board['rows']
                       if int(r['store_number']) == 105)
        assert row_105['priority_rank'] == 1

    def test_import_owner_actor_and_allowlist(self, fixture_stores, client):
        # Owner download -> edit -> upload. The owner export is ALREADY the
        # gap board; the POST must pass the fail-closed allowlist and audit
        # as 'owner-xlsx'. Applying rank 1 to store 101 resequences 105.
        blob = _download_xlsx(client, headers={'X-View': 'owner'})
        header, rows = _sheet_rows(blob)
        assert {r[header.index('Store #')] for r in rows} == GAP_STORES
        edited = _build_xlsx(['Store #', 'Rank'], [[101, 1]])
        resp = _post_xlsx(client, edited, headers={'X-View': 'owner'})
        assert resp.status_code == 200, resp.get_data(as_text=True)
        body = resp.get_json()
        assert body['updated'] == 1
        assert body['changed_by'] == 'owner-xlsx'
        state = _rank_state()
        assert state[101] == 1 and state[105] == 2
        _assert_unique_sequential(state)
        audits = _history_rows('priority_rank', 101)
        assert audits[-1]['changed_by'] == 'owner-xlsx'

    def test_import_skip_reasons(self, fixture_stores, client):
        edited = _build_xlsx(
            ['Store #', 'Rank', 'Owner Status'],
            [
                [999999, 5, None],              # unknown store
                ['abc', 5, None],               # unparseable store number
                [102, 'abc', None],             # bad rank
                [106, 0, None],                 # rank must be >= 1
                [103, 3, 'weird_status'],       # rank ok, status invalid
                [104, None, 'completed'],       # status change applies
                [104, None, 'completed'],       # duplicate row
            ])
        resp = _post_xlsx(client, edited)
        assert resp.status_code == 200, resp.get_data(as_text=True)
        body = resp.get_json()
        reasons = {(s['store_number'], s['reason'].split(',')[0].split(' got')[0])
                   for s in body['skipped']}
        assert (999999, 'not in the territory book') in reasons
        assert ('abc', 'bad store number') in reasons
        assert any(s['store_number'] == 102 and 'bad rank' in s['reason']
                   for s in body['skipped'])
        assert any(s['store_number'] == 106 and 'rank must be >= 1' in s['reason']
                   for s in body['skipped'])
        assert any(s['store_number'] == 103 and 'invalid owner_status' in s['reason']
                   for s in body['skipped'])
        assert any(s['store_number'] == 104 and s['reason'] == 'duplicate row'
                   for s in body['skipped'])
        assert len(body['skipped']) == 6
        # The two valid changes still applied: 103 -> rank 3, 104 -> completed.
        assert body['updated'] == 2
        state = _rank_state()
        assert state[103] == 3
        _assert_unique_sequential(state)
        db = _db()
        row = db.execute(
            "SELECT owner_status, owner_status_updated_at FROM territory_stores "
            "WHERE store_number=?", (104,)).fetchone()
        db.close()
        assert row['owner_status'] == 'completed'
        assert row['owner_status_updated_at'] is not None
        audits = _history_rows('owner_status', 104)
        assert audits[-1]['changed_by'] == 'internal-xlsx'
        assert audits[-1]['old_value'] == 'order_received'
        assert audits[-1]['new_value'] == 'completed'

    def test_import_rejects_oversize_file(self, fixture_stores, client):
        resp = _post_xlsx(client, b'x' * (1024 * 1024 + 1))
        assert resp.status_code == 413

    def test_import_rejects_garbage_bytes(self, fixture_stores, client):
        resp = _post_xlsx(client, b'this is not a workbook')
        assert resp.status_code == 400

    def test_import_requires_file_field(self, fixture_stores, client):
        resp = client.post('/api/top100/import-xlsx', data={},
                           content_type='multipart/form-data')
        assert resp.status_code == 400

    def test_import_requires_known_columns(self, fixture_stores, client):
        resp = _post_xlsx(client, _build_xlsx(['Foo', 'Bar'], [[1, 2]]))
        assert resp.status_code == 400


class TestContactsHistory:
    """Section 2c: audited contact fields + the restore endpoint."""

    def test_store_update_audits_the_three_log_fields(self, fixture_stores, client, crm_store_id):
        resp = client.put(f'/api/stores/{crm_store_id}', json={
            'manager_name': 'Raj', 'asst_manager_name': 'Priya',
            'spirits_ambassador': 'Sam', 'updated_by': 'Namit'})
        assert resp.status_code == 200, resp.get_data(as_text=True)
        assert resp.get_json()['audited_fields'] == 3
        for field, val in (('manager_name', 'Raj'),
                           ('asst_manager_name', 'Priya'),
                           ('spirits_ambassador', 'Sam')):
            rows = _history_rows(f'contact:{field}', CRM_STORE_SN)
            assert len(rows) == 1
            assert rows[0]['old_value'] == ''
            assert rows[0]['new_value'] == val
            assert rows[0]['changed_by'] == 'Namit'

    def test_contacts_history_endpoint_shape(self, fixture_stores, client, crm_store_id):
        # Second change so manager_name has a real restore candidate.
        client.put(f'/api/stores/{crm_store_id}', json={
            'manager_name': 'Anita', 'updated_by': 'Ikshit'})
        resp = client.get(f'/api/crm/store/{CRM_STORE_SN}/contacts/history')
        assert resp.status_code == 200, resp.get_data(as_text=True)
        body = resp.get_json()
        assert body['store_number'] == CRM_STORE_SN
        assert body['store_id'] == crm_store_id
        assert set(body['contacts']) == {
            'manager_name', 'asst_manager_name', 'spirits_ambassador'}
        mgr = body['contacts']['manager_name']
        assert mgr['current'] == 'Anita'
        assert mgr['previous'] == 'Raj'
        assert len(mgr['history']) == 2
        newest = mgr['history'][0]
        assert newest['old'] == 'Raj' and newest['new'] == 'Anita'
        assert newest['changed_by'] == 'Ikshit'
        assert newest['changed_at']
        asst = body['contacts']['asst_manager_name']
        assert asst['current'] == 'Priya'
        assert asst['previous'] == ''       # only prior value was blank
        assert len(asst['history']) == 1

    def test_restore_is_a_normal_audited_update(self, fixture_stores, client, crm_store_id):
        # Restore = PUT the old value back through the same endpoint.
        resp = client.put(f'/api/stores/{crm_store_id}', json={
            'manager_name': 'Raj', 'updated_by': 'Namit'})
        assert resp.status_code == 200
        assert resp.get_json()['audited_fields'] == 1
        body = client.get(
            f'/api/crm/store/{CRM_STORE_SN}/contacts/history').get_json()
        mgr = body['contacts']['manager_name']
        assert mgr['current'] == 'Raj'
        assert mgr['previous'] == 'Anita'
        assert len(mgr['history']) == 3
        assert mgr['history'][0]['old'] == 'Anita'
        assert mgr['history'][0]['new'] == 'Raj'

    def test_blanking_keeps_prior_value_one_tap_away(self, fixture_stores, client, crm_store_id):
        client.put(f'/api/stores/{crm_store_id}', json={
            'manager_name': '', 'updated_by': 'Namit'})
        body = client.get(
            f'/api/crm/store/{CRM_STORE_SN}/contacts/history').get_json()
        mgr = body['contacts']['manager_name']
        assert mgr['current'] == ''
        assert mgr['previous'] == 'Raj'     # nothing is ever hard-deleted
        assert mgr['history'][0]['new'] == ''

    def test_history_caps_at_ten_prior_values(self, fixture_stores, client, crm_store_id):
        for i in range(1, 13):
            client.put(f'/api/stores/{crm_store_id}', json={
                'spirits_ambassador': f'Ambassador {i}', 'updated_by': 'Ed'})
        body = client.get(
            f'/api/crm/store/{CRM_STORE_SN}/contacts/history').get_json()
        amb = body['contacts']['spirits_ambassador']
        assert amb['current'] == 'Ambassador 12'
        assert len(amb['history']) == 10
        assert amb['history'][0]['new'] == 'Ambassador 12'
        assert amb['previous'] == 'Ambassador 11'

    def test_unknown_store_is_404(self, fixture_stores, client):
        resp = client.get('/api/crm/store/424242/contacts/history')
        assert resp.status_code == 404

    def test_owner_view_gets_403_fail_closed(self, fixture_stores, client):
        # Contact identities never reach the owner: the endpoint is absent
        # from the allowlist, so the fail-closed hook rejects both spellings.
        for hdr, qs in (({'X-View': 'owner'}, ''), ({}, '?view=owner')):
            resp = client.get(
                f'/api/crm/store/{CRM_STORE_SN}/contacts/history{qs}',
                headers=hdr)
            assert resp.status_code == 403
            assert resp.get_json() == {'error': 'owner view: not permitted'}
