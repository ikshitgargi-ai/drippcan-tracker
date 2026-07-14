"""Dripp Tracker feature tests — territory book, live lcbo.com engine,
3-way reconciliation, and the changes feed.

Tracking accuracy IS the product: these tests assert that
  - territory ingest is idempotent and never destroys manual edits,
  - discovery excludes stores already in the territory book,
  - live snapshots are APPEND-ONLY (a new batch never touches old rows),
  - reconciliation produces each of the 6 flags correctly,
  - listing changes are tagged tier + route_day + attribution on both
    sides of LAUNCH_DATE,
  - no COALESCE(<date col>, '') patterns exist (prod-only crash we hit).

Run with: python3 -m pytest tests/test_dripp.py -v
"""
import json
import os
import re
import sqlite3
import sys

import pytest

# Force SQLite for tests so we never touch production Postgres.
os.environ.pop('DATABASE_URL', None)
os.environ.pop('ADMIN_TOKEN', None)
os.environ.pop('API_KEY', None)
os.environ.pop('SOD_CRON_TOKEN', None)

TEST_DB_DIR = '/tmp/drippcan_dripp_test'
os.makedirs(TEST_DB_DIR, exist_ok=True)
TEST_DB = os.path.join(TEST_DB_DIR, 'drippcan.db')

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

PHOENIX = '0014318'
DAYAA = '0044451'


@pytest.fixture(scope='module')
def app_module():
    """Import app.py fresh against an isolated SQLite file.

    DB_DIR is re-asserted here (not only at module import) because pytest
    imports every test module during collection and the LAST module import
    wins the env var. Restored on teardown so other test modules are
    unaffected.
    """
    prev_db_dir = os.environ.get('DB_DIR')
    os.environ['DB_DIR'] = TEST_DB_DIR
    if os.path.exists(TEST_DB):
        os.remove(TEST_DB)
    for mod in list(sys.modules):
        if mod == 'app' or mod.startswith('app.'):
            del sys.modules[mod]
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        'app', os.path.join(os.path.dirname(__file__), '..', 'app.py'))
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    yield m
    if prev_db_dir is None:
        os.environ.pop('DB_DIR', None)
    else:
        os.environ['DB_DIR'] = prev_db_dir


@pytest.fixture
def client(app_module):
    app_module.app.config['TESTING'] = True
    app_module._rate_buckets.clear()  # tests fire faster than 50 req/s
    with app_module.app.test_client() as c:
        yield c


REP_NAME_RE = re.compile(r'\b(Ikshit|Vaneet|Ed|Namit)\b')


def _db():
    conn = sqlite3.connect(TEST_DB)
    conn.row_factory = sqlite3.Row
    return conn


@pytest.fixture(scope='module')
def ingested(app_module):
    """Run the territory ingest once for the module (later tests need it)."""
    app_module.app.config['TESTING'] = True
    with app_module.app.test_client() as c:
        resp = c.post('/api/territory/ingest')
        assert resp.status_code == 200, resp.get_data(as_text=True)
        return resp.get_json()


def _fake_scraper(per_sku):
    """Build a _live_scrape_sku stand-in from {padded_sku: [(store, qty), ...]}."""
    def fake(sku):
        rows = per_sku.get(str(sku).zfill(7))
        if rows is None:
            return [], 'no store rows parsed — product may be delisted or page layout changed'
        return ([
            {'store_number': str(sn), 'city': 'Toronto', 'intersection': 'Test & Test',
             'store_name': f'Test Store {sn}', 'address': f'{sn} Test St', 'phone': '',
             'quantity': qty}
            for sn, qty in rows
        ], None)
    return fake


# ---------------------------------------------------------------------------
# TERRITORY ingest
# ---------------------------------------------------------------------------

class TestTerritoryIngest:
    def test_first_ingest_loads_189_rows(self, ingested, client):
        assert ingested['inserted'] == 189
        assert ingested['total'] == 189
        db = _db()
        assert db.execute('SELECT COUNT(*) FROM territory_stores').fetchone()[0] == 189
        assert db.execute(
            "SELECT COUNT(*) FROM territory_stores WHERE tier='routed'").fetchone()[0] == 73
        db.close()

    def test_default_priority_ranks_routed_by_route_sequence(self, ingested):
        db = _db()
        routed = db.execute(
            "SELECT store_number, route_day, route_stop, priority_rank "
            "FROM territory_stores WHERE tier='routed'").fetchall()
        assert all(r['priority_rank'] is not None for r in routed)
        ranks = [r['priority_rank'] for r in routed]
        assert len(set(ranks)) == 73  # unique
        # Rank order must follow (route_day, route_stop)
        by_rank = sorted(routed, key=lambda r: r['priority_rank'])
        seq = [(r['route_day'], r['route_stop']) for r in by_rank]
        assert seq == sorted(seq)
        assert by_rank[0]['priority_rank'] == 1
        db.close()

    def test_reingest_is_idempotent_and_preserves_manual_edits(self, ingested, client):
        db = _db()
        sn = db.execute(
            "SELECT store_number FROM territory_stores WHERE tier='routed' LIMIT 1"
        ).fetchone()[0]
        db.execute(
            "UPDATE territory_stores SET priority_rank=555, owner_status='order_received' "
            "WHERE store_number=?", (sn,))
        db.commit()
        db.close()

        resp = client.post('/api/territory/ingest')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['inserted'] == 0
        assert body['updated'] == 189
        assert body['total'] == 189  # re-run safe: nothing deleted, nothing duplicated

        db = _db()
        row = db.execute(
            "SELECT priority_rank, owner_status FROM territory_stores WHERE store_number=?",
            (sn,)).fetchone()
        assert row['priority_rank'] == 555
        assert row['owner_status'] == 'order_received'
        db.close()

    def test_rank_assignments_are_audited(self, ingested):
        db = _db()
        n = db.execute(
            "SELECT COUNT(*) FROM territory_status_history "
            "WHERE field='priority_rank' AND changed_by='seed-default'").fetchone()[0]
        assert n >= 73  # every default rank assignment landed in the audit trail
        db.close()


class TestTerritoryGet:
    def test_territory_list_and_filters(self, ingested, client):
        body = client.get('/api/territory?nocache=1').get_json()
        assert body['count'] == 189
        first = body['stores'][0]
        assert first['tier'] == 'routed'  # routed stores sort first
        for key in ('sku_presence', 'last_touchpoint', 'lat', 'lng'):
            assert key in first
        assert set(first['sku_presence'].keys()) == {PHOENIX, DAYAA}

        routed = client.get('/api/territory?tier=routed&nocache=1').get_json()
        assert routed['count'] == 73
        brampton = client.get('/api/territory?city=Brampton&nocache=1').get_json()
        assert brampton['count'] >= 1
        assert all(s['city'].lower() == 'brampton' for s in brampton['stores'])

    def test_territory_never_500_on_garbage_params(self, ingested, client):
        resp = client.get('/api/territory?tier=zzz&q=%25%27--&nocache=1')
        assert resp.status_code == 200
        assert resp.get_json()['count'] == 0


# ---------------------------------------------------------------------------
# Discovery
# ---------------------------------------------------------------------------

class TestDiscovery:
    def test_discovery_excludes_existing_and_non_gta(self, ingested, client):
        db = _db()
        db.execute(
            "INSERT OR IGNORE INTO stores (store_number, account, address, city, postal) "
            "VALUES (9901, 'LCBO #9901 Test GTA', '1 Yonge St', 'Toronto', 'M1M1M1')")
        db.execute(
            "INSERT OR IGNORE INTO stores (store_number, account, address, city, postal) "
            "VALUES (9902, 'LCBO #9902 Test Non-GTA', '1 Bank St', 'Ottawa', 'K1K1K1')")
        terr_sn = db.execute(
            "SELECT store_number FROM territory_stores LIMIT 1").fetchone()[0]
        db.commit()
        db.close()

        body = client.get('/api/territory/discovery?nocache=1').get_json()
        numbers = {c['store_number'] for c in body['candidates']}
        assert 9901 in numbers          # GTA store not yet in territory
        assert 9902 not in numbers      # non-GTA city excluded
        assert terr_sn not in numbers   # already in the territory book

    def test_discovery_add_moves_store_into_territory(self, ingested, client):
        resp = client.post('/api/territory/discovery/add',
                           json={'store_number': 9901, 'added_by': 'Ikshit'})
        assert resp.status_code == 200
        assert resp.get_json()['status'] == 'added'

        db = _db()
        row = db.execute(
            "SELECT tier, source, city FROM territory_stores WHERE store_number=9901"
        ).fetchone()
        assert row['tier'] == 'discovered'
        assert row['source'] == 'discovery'
        assert row['city'] == 'Toronto'
        audit = db.execute(
            "SELECT COUNT(*) FROM territory_status_history WHERE store_number=9901"
        ).fetchone()[0]
        assert audit >= 1
        db.close()

        body = client.get('/api/territory/discovery?nocache=1').get_json()
        assert 9901 not in {c['store_number'] for c in body['candidates']}

        # Adding again is a harmless no-op (never deletes / never duplicates)
        again = client.post('/api/territory/discovery/add', json={'store_number': 9901})
        assert again.get_json()['status'] == 'exists'


# ---------------------------------------------------------------------------
# LIVE lcbo.com engine — append-only snapshots + events
# ---------------------------------------------------------------------------

BATCH1 = {PHOENIX: [(1, 6), (2, 10), (3, 4)], DAYAA: [(1, 5)]}
BATCH2 = {PHOENIX: [(1, 6), (2, 15), (4, 8)], DAYAA: [(1, 5)]}


class TestLiveEngine:
    def test_two_batches_are_append_only(self, ingested, app_module, client, monkeypatch):
        monkeypatch.setattr(app_module, 'LIVE_SCRAPE_GAP_SECONDS', 0)
        monkeypatch.setattr(app_module, '_live_scrape_sku', _fake_scraper(BATCH1))
        s1 = app_module.run_live_batch(triggered_by='test')
        assert s1['status'] == 'ok'
        assert s1['row_count'] == 4

        # Second batch via the on-demand endpoint (admin: localhost dev mode)
        monkeypatch.setattr(app_module, '_live_scrape_sku', _fake_scraper(BATCH2))
        resp = client.post('/api/live/refresh')
        assert resp.status_code == 200
        s2 = resp.get_json()
        assert s2['status'] == 'ok'

        db = _db()
        total = db.execute('SELECT COUNT(*) FROM lcbo_live_snapshots').fetchone()[0]
        assert total == 8  # 4 + 4: batch 2 did NOT update/delete batch 1 rows
        batches = db.execute(
            "SELECT COUNT(DISTINCT batch_id) FROM lcbo_live_snapshots").fetchone()[0]
        assert batches == 2
        # Batch 1 rows still intact and unchanged
        old_qty = db.execute(
            "SELECT qty FROM lcbo_live_snapshots WHERE sku=? AND store_number=2 AND batch_id=?",
            (PHOENIX, s1['batch_id'])).fetchone()[0]
        assert old_qty == 10
        db.close()

    def test_live_listing_events_detected_between_batches(self, app_module):
        db = _db()
        events = {
            (r['event_type'], r['store_number']): r
            for r in db.execute(
                "SELECT event_type, store_number, old_qty, new_qty "
                "FROM live_listing_events WHERE sku=?", (PHOENIX,)).fetchall()
        }
        db.close()
        assert ('LIVE_NEW_LISTING', 4) in events        # store appeared
        assert events[('LIVE_NEW_LISTING', 4)]['new_qty'] == 8
        assert ('LIVE_RESTOCK', 2) in events            # 10 -> 15
        assert events[('LIVE_RESTOCK', 2)]['old_qty'] == 10
        assert events[('LIVE_RESTOCK', 2)]['new_qty'] == 15
        assert ('LIVE_DELISTED', 3) in events           # store disappeared
        # Unchanged store never generates noise
        assert ('LIVE_RESTOCK', 1) not in events

    def test_live_batches_fold_into_listing_ledger(self, app_module):
        # The canonical ledger must capture the live signal from the two
        # batches above: LISTED for the newly seen store, DELISTED for the one
        # that vanished, RECONFIRMED (deduped per day) for every store seen.
        db = _db()
        rows = {
            (r['event'], r['store_number'])
            for r in db.execute(
                "SELECT event, store_number FROM listing_ledger "
                "WHERE sku=? AND source='live'", (PHOENIX,)).fetchall()
        }
        total_live = db.execute(
            "SELECT COUNT(*) FROM listing_ledger WHERE source='live'").fetchone()[0]
        db.close()
        assert ('LISTED', 4) in rows          # LIVE_NEW_LISTING folded
        assert ('DELISTED', 3) in rows        # LIVE_DELISTED folded
        assert ('RECONFIRMED', 1) in rows and ('RECONFIRMED', 2) in rows
        # Phoenix 5 (RECONFIRMED 1/2/3 + LISTED 4 + DELISTED 3) + Dayaa 1;
        # batch 2 re-sightings dedupe to no-ops (one row per source per day).
        assert total_live == 6

    def test_live_latest_returns_only_newest_batch(self, client):
        body = client.get(f'/api/live/latest?sku={PHOENIX}&nocache=1').get_json()
        sku_block = body['skus'][PHOENIX]
        by_store = {s['store_number']: s['qty'] for s in sku_block['stores']}
        assert by_store == {1: 6, 2: 15, 4: 8}  # batch 2 view; store 3 gone
        assert sku_block['total_units'] == 29
        assert sku_block['checked_at']

    def test_live_store_time_series(self, client):
        body = client.get(f'/api/live/store/2?sku={PHOENIX}&nocache=1').get_json()
        qtys = [p['qty'] for p in body['series']]
        assert qtys == [10, 15]  # both batches preserved, oldest first

    def test_failed_scrape_records_error_and_touches_nothing(self, app_module, monkeypatch):
        db = _db()
        before = db.execute('SELECT COUNT(*) FROM lcbo_live_snapshots').fetchone()[0]
        db.close()

        monkeypatch.setattr(app_module, 'LIVE_SCRAPE_GAP_SECONDS', 0)
        monkeypatch.setattr(app_module, '_live_scrape_sku',
                            lambda sku: ([], 'scrape error: boom'))
        summary = app_module.run_live_batch(triggered_by='test')
        assert summary['status'] == 'error'

        db = _db()
        after = db.execute('SELECT COUNT(*) FROM lcbo_live_snapshots').fetchone()[0]
        assert after == before  # append-only: failure adds nothing, changes nothing
        batch = db.execute(
            "SELECT status, error FROM lcbo_live_batches WHERE batch_id=?",
            (summary['batch_id'],)).fetchone()
        assert batch['status'] == 'error'
        assert 'boom' in batch['error']
        db.close()

    def test_batch_dedupes_duplicate_store_rows(self, ingested, app_module, monkeypatch):
        """lcbo.com renders each store block twice on the storeinventory page
        (verified live 2026-07-14) — one snapshot row per (sku, store) per
        batch, or every inventory count doubles."""
        dup_rows = {PHOENIX: [(901, 4), (901, 4), (902, 2)], DAYAA: [(901, 3), (901, 3)]}
        monkeypatch.setattr(app_module, 'LIVE_SCRAPE_GAP_SECONDS', 0)
        monkeypatch.setattr(app_module, '_live_scrape_sku', _fake_scraper(dup_rows))
        summary = app_module.run_live_batch(triggered_by='test')
        assert summary['status'] == 'ok'
        assert summary['row_count'] == 3  # 901+902 Phoenix, 901 Dayaa

        db = _db()
        rows = db.execute(
            "SELECT sku, store_number, COUNT(*) c FROM lcbo_live_snapshots "
            "WHERE batch_id=? GROUP BY sku, store_number",
            (summary['batch_id'],)).fetchall()
        db.close()
        assert len(rows) == 3
        assert all(r['c'] == 1 for r in rows)


# ---------------------------------------------------------------------------
# 3-way reconciliation — all six flags
# ---------------------------------------------------------------------------

class TestReconcile:
    @pytest.fixture(scope='class')
    def scenario(self, ingested, app_module):
        """Six territory stores, one per flag state, for Phoenix."""
        db = _db()
        stores = [r[0] for r in db.execute(
            "SELECT store_number FROM territory_stores WHERE active=1 "
            "ORDER BY store_number LIMIT 6").fetchall()]
        t_match, t_sod_lags, t_live_lags, t_rep, t_no_sod, t_no_live = stores
        today = app_module._toronto_today().isoformat()
        # SOD latest snapshot (no row for t_no_sod)
        for sn, on_hand in ((t_match, 6), (t_sod_lags, 2), (t_live_lags, 9),
                            (t_rep, 10), (t_no_live, 7)):
            db.execute(
                "INSERT OR IGNORE INTO sod_inventory "
                "(sku, store_number, snapshot_date, status, on_hand, product_name) "
                "VALUES (?,?,?,?,?,?)",
                (PHOENIX, sn, today, 'L', on_hand, 'PHOENIX ULTRA SMOOTH VODKA'))
        # Rep observation: says 2 units where both systems say 10 (delta 8 >= 3)
        db.execute(
            "INSERT INTO rep_listing_observations (sku, store_number, rep, on_shelf, units) "
            "VALUES (?,?,?,?,?)", (PHOENIX, t_rep, 'Ikshit', 1, 2))
        db.commit()
        db.close()
        return {'match': t_match, 'sod_lags': t_sod_lags, 'live_lags': t_live_lags,
                'rep': t_rep, 'no_sod': t_no_sod, 'no_live': t_no_live}

    def test_all_six_flags(self, scenario, app_module, client, monkeypatch):
        # Live batch covering the scenario (no row for t_no_live)
        live_rows = {PHOENIX: [
            (scenario['match'], 6), (scenario['sod_lags'], 8),
            (scenario['live_lags'], 4), (scenario['rep'], 10),
            (scenario['no_sod'], 5),
        ], DAYAA: [(scenario['match'], 3)]}
        monkeypatch.setattr(app_module, 'LIVE_SCRAPE_GAP_SECONDS', 0)
        monkeypatch.setattr(app_module, '_live_scrape_sku', _fake_scraper(live_rows))
        assert app_module.run_live_batch(triggered_by='test')['status'] == 'ok'

        body = client.get(f'/api/reconcile?days=7&sku={PHOENIX}&nocache=1').get_json()
        flags = {r['store_number']: r for r in body['rows']}
        assert flags[scenario['match']]['flag'] == 'MATCH'
        assert flags[scenario['sod_lags']]['flag'] == 'SOD_LAGS_LIVE'
        assert flags[scenario['live_lags']]['flag'] == 'LIVE_LAGS_SOD'
        assert flags[scenario['rep']]['flag'] == 'REP_MISMATCH'
        assert flags[scenario['no_sod']]['flag'] == 'MISSING_FROM_SOD'
        assert flags[scenario['no_live']]['flag'] == 'MISSING_FROM_LIVE'

        # A diff is never hidden: raw values + deltas + timestamps ride along
        rep_row = flags[scenario['rep']]
        assert rep_row['sod_on_hand'] == 10 and rep_row['live_qty'] == 10
        assert rep_row['rep_units'] == 2 and rep_row['delta_rep_live'] == -8
        assert rep_row['rep_observed_at'] and rep_row['live_checked_at']
        assert rep_row['sod_snapshot_date']
        lag_row = flags[scenario['sod_lags']]
        assert lag_row['delta_sod_live'] == -6

        # Every source's last-checked timestamp is surfaced
        src = body['sources'][PHOENIX]
        assert src['sod_latest_snapshot'] and src['live_checked_at']
        # Summary counts every flag
        assert sum(body['summary'].values()) == len(body['rows'])

    def test_reconcile_never_500_on_empty_data(self, app_module, client):
        resp = client.get(f'/api/reconcile?days=abc&sku=9999999&nocache=1')
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
# Changes feed + attribution on both sides of LAUNCH_DATE
# ---------------------------------------------------------------------------

class TestChangesAndAttribution:
    @pytest.fixture(scope='class')
    def change_stores(self, ingested, app_module):
        db = _db()
        stores = [r[0] for r in db.execute(
            "SELECT store_number FROM territory_stores WHERE tier='routed' "
            "ORDER BY store_number LIMIT 4").fetchall()]
        s_converted, s_organic, s_baseline, s_dropped = stores
        # New listings: two after LAUNCH_DATE (2026-07-15), one before
        for sn, d, ctype in ((s_converted, '2026-07-16', 'NEW_LISTING'),
                             (s_organic, '2026-07-16', 'NEW_LISTING'),
                             (s_baseline, '2026-07-10', 'NEW_LISTING'),
                             (s_dropped, '2026-07-12', 'DROPPED')):
            db.execute(
                "INSERT OR IGNORE INTO sod_store_sku_changes "
                "(sku, store_number, change_date, old_status, new_status, change_type) "
                "VALUES (?,?,?,?,?,?)", (PHOENIX, sn, d, None, 'L', ctype))
        # Touchpoints: visit BEFORE the listing at s_converted and s_baseline
        rep_id = db.execute(
            "SELECT id FROM reps WHERE name='Ikshit'").fetchone()[0]
        for sn, ts in ((s_converted, '2026-07-15 09:00:00'),
                       (s_baseline, '2026-07-01 09:00:00')):
            store_row = db.execute(
                "SELECT id FROM stores WHERE store_number=?", (sn,)).fetchone()
            if store_row is None:
                db.execute("INSERT INTO stores (store_number, account) VALUES (?,?)",
                           (sn, f'LCBO #{sn}'))
                store_id = db.execute(
                    "SELECT id FROM stores WHERE store_number=?", (sn,)).fetchone()[0]
            else:
                store_id = store_row[0]
            db.execute(
                "INSERT INTO activities (store_id, rep_id, activity_type, created_at) "
                "VALUES (?,?,?,?)", (store_id, rep_id, 'visit', ts))
        db.commit()
        db.close()
        return {'converted': s_converted, 'organic': s_organic,
                'baseline': s_baseline, 'dropped': s_dropped}

    def test_changes_tagged_with_tier_route_and_attribution(self, change_stores, client):
        body = client.get(f'/api/changes?days=14&sku={PHOENIX}&nocache=1').get_json()
        rows = {(r['store_number'], r['change_type']): r
                for r in body['rows'] if r['source'] == 'sod'}

        conv = rows[(change_stores['converted'], 'NEW_LISTING')]
        assert conv['attribution'] == 'rep_converted'   # touchpoint before listing
        assert conv['kind'] == 'new_listing'
        assert conv['in_territory'] is True
        assert conv['tier'] == 'routed'
        assert conv['route_day'] is not None

        org = rows[(change_stores['organic'], 'NEW_LISTING')]
        assert org['attribution'] == 'organic'          # no touchpoint at this store

        base = rows[(change_stores['baseline'], 'NEW_LISTING')]
        assert base['attribution'] == 'baseline'        # listed on/before LAUNCH_DATE
        assert base['attribution'] != 'rep_converted'   # pre-launch is never claimed

        drop = rows[(change_stores['dropped'], 'DROPPED')]
        assert drop['kind'] == 'delisting'
        assert drop['attribution'] is None              # attribution is for listings only

        s = body['summary']
        assert s['new_listings'] >= 3 and s['delistings'] >= 1
        assert s['rep_converted'] >= 1 and s['organic'] >= 1 and s['baseline'] >= 1

    def test_changes_includes_live_events(self, change_stores, client):
        body = client.get(f'/api/changes?days=14&sku={PHOENIX}&nocache=1').get_json()
        live = [r for r in body['rows'] if r['source'] == 'live']
        assert live, 'live_listing_events from earlier batches must surface in /api/changes'
        assert {r['change_type'] for r in live} >= {'LIVE_NEW_LISTING'}
        for r in live:
            assert r['kind'] in ('new_listing', 'delisting', 'restock')

    def test_changes_never_500_on_garbage_days(self, ingested, client):
        resp = client.get('/api/changes?days=banana&nocache=1')
        assert resp.status_code == 200
        assert resp.get_json()['days'] == 7  # falls back to the default window


# ---------------------------------------------------------------------------
# Hygiene: the prod-only date-COALESCE crash can never come back
# ---------------------------------------------------------------------------

class TestHygiene:
    def test_no_coalesce_on_date_columns(self):
        """Postgres: COALESCE(date_col, '') crashes (date vs text). CAST first."""
        src = open(os.path.join(os.path.dirname(__file__), '..', 'app.py')).read()
        bad = re.findall(
            r"COALESCE\(\s*[A-Za-z_.]*(?:date|_at)[A-Za-z_.]*\s*,\s*''", src)
        assert bad == [], f'date-column COALESCE with empty string found: {bad}'

    def test_live_row_regex_parses_reference_markup(self, app_module):
        html = (
            '<div><p class="city_txt">Toronto</p>\n'
            '<p class="name_txt">Queens Quay &amp; Cooper</p>\n'
            '<p class="address_txt">2 Cooper St.</p>\n'
            '<p class="quantity_avail_txt">12</p>\n'
            '<a href="https://www.lcbo.com/en/stores/toronto-queens-quay-217">store</a></div>'
        )
        m = app_module.LIVE_ROW_RE.findall(html)
        assert m == [('Toronto', 'Queens Quay &amp; Cooper', '2 Cooper St.', '12', '217')]


# ---------------------------------------------------------------------------
# Conversion scoreboard (/api/conversion) — rep_converted vs organic
# ---------------------------------------------------------------------------

def _ensure_store_row(db, sn):
    """Make sure a stores-table row exists for store_number sn; return its id."""
    row = db.execute("SELECT id FROM stores WHERE store_number=?", (sn,)).fetchone()
    if row is None:
        db.execute("INSERT INTO stores (store_number, account) VALUES (?,?)",
                   (sn, f'LCBO #{sn}'))
        row = db.execute("SELECT id FROM stores WHERE store_number=?", (sn,)).fetchone()
    return row[0]


class TestConversion:
    @pytest.fixture(scope='class')
    def conv_stores(self, ingested, app_module):
        """Two untouched high-numbered routed stores: one visited before its
        Dayaa listing (rep_converted), one never visited (organic)."""
        db = _db()
        stores = [r[0] for r in db.execute(
            "SELECT store_number FROM territory_stores WHERE tier='routed' "
            "ORDER BY store_number DESC LIMIT 2").fetchall()]
        s_conv, s_org = stores
        for sn, d in ((s_conv, '2026-07-17'), (s_org, '2026-07-17')):
            db.execute(
                "INSERT OR IGNORE INTO sod_store_sku_changes "
                "(sku, store_number, change_date, old_status, new_status, change_type) "
                "VALUES (?,?,?,?,?,?)", (DAYAA, sn, d, None, 'L', 'NEW_LISTING'))
        rep_id = db.execute("SELECT id FROM reps WHERE name='Ikshit'").fetchone()[0]
        store_id = _ensure_store_row(db, s_conv)
        db.execute(
            "INSERT INTO activities (store_id, rep_id, activity_type, created_at) "
            "VALUES (?,?,?,?)", (store_id, rep_id, 'visit', '2026-07-16 10:00:00'))
        db.commit()
        db.close()
        return {'conv': s_conv, 'org': s_org}

    def test_conversion_scoreboard(self, conv_stores, client):
        body = client.get(f'/api/conversion?days=60&sku={DAYAA}&nocache=1').get_json()
        assert body['launch_date'] == '2026-07-15'
        assert body['touchpoints'] >= 1
        assert body['stores_touched'] >= 1
        rows = {r['store_number']: r for r in body['per_store']}

        conv = rows[conv_stores['conv']]
        assert conv['attribution'] == 'rep_converted'  # visit 07-16 < listing 07-17
        assert conv['rep'] == 'Ikshit'
        assert conv['touch_description'].startswith('Ikshit visit on 2026-07-16')
        assert conv['tier'] == 'routed' and conv['in_territory'] is True

        org = rows[conv_stores['org']]
        assert org['attribution'] == 'organic'          # zero touchpoints ever
        assert org['rep'] is None

        # Counters are internally consistent and the rate is checkable math
        assert body['new_listings'] == len(body['per_store'])
        assert body['rep_converted'] >= 1 and body['organic'] >= 1
        convertible = body['rep_converted'] + body['organic']
        assert body['conversion_rate'] == round(
            100.0 * body['rep_converted'] / convertible, 1)

    def test_conversion_owner_safe(self, conv_stores, client):
        resp = client.get(f'/api/conversion?days=60&sku={DAYAA}&view=owner&nocache=1')
        assert resp.status_code == 200
        text = resp.get_data(as_text=True)
        assert not REP_NAME_RE.search(text), 'rep name leaked to owner view'
        rows = {r['store_number']: r for r in resp.get_json()['per_store']}
        conv = rows[conv_stores['conv']]
        # Owner mode v2: rep identity becomes the GTA region label
        assert conv['rep'] == 'GTA CENTRAL'
        assert conv['touch_description'].startswith('GTA CENTRAL visit on 2026-07-16')

    def test_conversion_never_500_on_garbage(self, ingested, client):
        resp = client.get('/api/conversion?days=banana&nocache=1')
        assert resp.status_code == 200
        assert resp.get_json()['days'] == 30


# ---------------------------------------------------------------------------
# TOP-100 priority board — ranking, the owner's two writes, ALWAYS audited
# ---------------------------------------------------------------------------

class TestTop100:
    def test_board_default_ranking(self, ingested, client):
        body = client.get('/api/top100?nocache=1').get_json()
        assert body['count'] == 100  # 190 territory stores -> capped at 100
        rows = body['rows']
        ranks = [r['priority_rank'] for r in rows]
        ranked = [x for x in ranks if x is not None]
        # Ranked stores come first and in ascending rank order
        assert ranks[:len(ranked)] == ranked
        assert ranked == sorted(ranked)
        assert ranked, 'seed-default ranks must populate the board'
        first = rows[0]
        for key in ('skus', 'last_touchpoint', 'owner_status', 'conversion'):
            assert key in first
        assert set(first['skus'].keys()) == {PHOENIX, DAYAA}
        for sku_block in first['skus'].values():
            for key in ('listed', 'on_hand', 'live_qty', 'conversion'):
                assert key in sku_block

    def test_priority_write_is_audited(self, ingested, client):
        body = client.get('/api/top100?nocache=1').get_json()
        sn = body['rows'][0]['store_number']
        resp = client.post('/api/top100/priority',
                           json={'store_number': sn, 'rank': 2, 'changed_by': 'Namit'})
        assert resp.status_code == 200
        assert resp.get_json()['changed_by'] == 'Namit'

        db = _db()
        row = db.execute(
            "SELECT priority_rank FROM territory_stores WHERE store_number=?",
            (sn,)).fetchone()
        assert row['priority_rank'] == 2
        audit = db.execute(
            "SELECT old_value, new_value, changed_by FROM territory_status_history "
            "WHERE store_number=? AND field='priority_rank' "
            "ORDER BY id DESC LIMIT 1", (sn,)).fetchone()
        assert audit['new_value'] == '2'
        assert audit['changed_by'] == 'Namit'
        db.close()

    def test_owner_priority_write_audits_as_owner(self, ingested, client):
        body = client.get('/api/top100?nocache=1').get_json()
        sn = body['rows'][1]['store_number']
        resp = client.post('/api/top100/priority',
                           json={'store_number': sn, 'rank': 7},
                           headers={'X-View': 'owner'})
        assert resp.status_code == 200
        assert resp.get_json()['changed_by'] == 'owner'
        db = _db()
        audit = db.execute(
            "SELECT changed_by FROM territory_status_history "
            "WHERE store_number=? AND field='priority_rank' "
            "ORDER BY id DESC LIMIT 1", (sn,)).fetchone()
        assert audit['changed_by'] == 'owner'
        db.close()

    def test_status_write_is_audited(self, ingested, client):
        body = client.get('/api/top100?nocache=1').get_json()
        sn = body['rows'][0]['store_number']
        resp = client.post('/api/top100/status',
                           json={'store_number': sn,
                                 'owner_status': 'listing_received',
                                 'note': 'send more stock',
                                 'changed_by': 'Vaneet'})
        assert resp.status_code == 200
        db = _db()
        row = db.execute(
            "SELECT owner_status, owner_status_note, owner_status_updated_at "
            "FROM territory_stores WHERE store_number=?", (sn,)).fetchone()
        assert row['owner_status'] == 'listing_received'
        assert row['owner_status_note'] == 'send more stock'
        assert row['owner_status_updated_at'] is not None
        audits = {r['field']: r for r in db.execute(
            "SELECT field, new_value, changed_by FROM territory_status_history "
            "WHERE store_number=? AND field IN ('owner_status','owner_status_note') "
            "ORDER BY id DESC LIMIT 2", (sn,)).fetchall()}
        assert audits['owner_status']['new_value'] == 'listing_received'
        assert audits['owner_status']['changed_by'] == 'Vaneet'
        assert audits['owner_status_note']['new_value'] == 'send more stock'
        db.close()

    def test_bad_writes_rejected(self, ingested, client):
        assert client.post('/api/top100/status',
                           json={'store_number': 1, 'owner_status': 'banana'}
                           ).status_code == 400
        assert client.post('/api/top100/priority',
                           json={'store_number': 1, 'rank': 'abc'}
                           ).status_code == 400
        assert client.post('/api/top100/priority',
                           json={'store_number': 999999, 'rank': 5}
                           ).status_code == 404

    def test_funnel_counts_the_board(self, ingested, client):
        body = client.get('/api/top100/funnel?nocache=1').get_json()
        assert body['board_size'] == 100
        assert sum(body['funnel'].values()) == body['board_size']
        assert body['funnel']['listing_received'] >= 1  # set in the status test
        assert set(body['funnel'].keys()) == {
            'none', 'listing_received', 'order_received', 'completed'}


# ---------------------------------------------------------------------------
# OWNER MODE — server-side anonymization on every owner surface
# ---------------------------------------------------------------------------

class TestOwnerMode:
    def test_internal_view_shows_rep_names(self, ingested, client):
        """Prove the owner assertions are not vacuous: internal reconcile
        carries the observing rep's real name."""
        text = client.get(f'/api/reconcile?days=7&sku={PHOENIX}&nocache=1'
                          ).get_data(as_text=True)
        assert 'Ikshit' in text

    @pytest.mark.parametrize('path', [
        '/api/territory?view=owner&nocache=1',
        f'/api/reconcile?days=7&sku={PHOENIX}&view=owner&nocache=1',
        '/api/changes?days=60&view=owner&nocache=1',
        '/api/conversion?days=60&view=owner&nocache=1',
        '/api/top100?view=owner&nocache=1',
        '/api/top100/funnel?view=owner&nocache=1',
    ])
    def test_owner_json_has_no_rep_names(self, ingested, path, client):
        resp = client.get(path)
        assert resp.status_code == 200
        text = resp.get_data(as_text=True)
        assert not REP_NAME_RE.search(text), f'rep name leaked in {path}'

    def test_owner_header_works_like_query_param(self, ingested, client):
        resp = client.get('/api/top100?nocache=1', headers={'X-View': 'owner'})
        assert resp.status_code == 200
        assert not REP_NAME_RE.search(resp.get_data(as_text=True))

    def test_owner_territory_strips_identity_but_keeps_touch_facts(
            self, ingested, client):
        body = client.get('/api/territory?view=owner&nocache=1').get_json()
        touched = [s for s in body['stores'] if s['last_touchpoint']]
        assert touched, 'expected at least one store with a touchpoint'
        for s in touched:
            # v2: identity becomes a GTA region label, never a name
            assert s['last_touchpoint']['rep'].startswith('GTA')
            assert s['last_touchpoint']['activity_type']  # type+date stay visible
            assert s['last_touchpoint']['created_at']
        for s in body['stores']:
            assert s.get('phone', '') == ''  # store contact details stay internal

    def test_owner_cache_never_leaks_across_views(self, ingested, client):
        """Warm the cache internally, then read the SAME url as owner —
        the owner copy must still be scrubbed (sanitize sits outside cache)."""
        internal = client.get(f'/api/reconcile?days=7&sku={PHOENIX}'
                              ).get_data(as_text=True)
        assert 'Ikshit' in internal
        owner = client.get(f'/api/reconcile?days=7&sku={PHOENIX}&view=owner'
                           ).get_data(as_text=True)
        assert not REP_NAME_RE.search(owner)


# ---------------------------------------------------------------------------
# EXPORTS — every board downloads as a real .xlsx workbook
# ---------------------------------------------------------------------------

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def _load_workbook(resp):
    import io as _io
    from openpyxl import load_workbook
    return load_workbook(_io.BytesIO(resp.data))


class TestExports:
    @pytest.mark.parametrize('path,min_data_rows', [
        ('/api/export/top100.xlsx', 100),
        ('/api/export/territory.xlsx', 189),
        ('/api/export/changes.xlsx?days=60', 1),
        ('/api/export/reconcile.xlsx?days=7', 1),
        ('/api/export/visits.xlsx', 1),
    ])
    def test_xlsx_export_is_a_valid_workbook(self, ingested, path,
                                             min_data_rows, client):
        resp = client.get(path)
        assert resp.status_code == 200
        assert resp.mimetype == XLSX_MIME
        assert 'attachment' in resp.headers.get('Content-Disposition', '')
        assert '.xlsx' in resp.headers.get('Content-Disposition', '')
        wb = _load_workbook(resp)
        ws = wb.active
        assert ws.max_row >= 1 + min_data_rows  # header + data
        headers = [c.value for c in ws[1]]
        assert headers and all(h for h in headers)

    def test_visits_export_internal_vs_owner(self, ingested, client):
        # Plant an activity with an internal note to prove notes stay internal
        db = _db()
        rep_id = db.execute("SELECT id FROM reps WHERE name='Namit'").fetchone()[0]
        sn = db.execute("SELECT store_number FROM territory_stores "
                        "ORDER BY store_number DESC LIMIT 1").fetchone()[0]
        store_id = _ensure_store_row(db, sn)
        db.execute(
            "INSERT INTO activities (store_id, rep_id, activity_type, notes, created_at) "
            "VALUES (?,?,?,?,?)",
            (store_id, rep_id, 'visit', 'Secret shelf intel about the manager',
             '2026-07-13 12:00:00'))
        db.commit()
        db.close()

        def cells_text(resp):
            ws = _load_workbook(resp).active
            return ' | '.join(str(c.value) for row in ws.iter_rows()
                              for c in row if c.value is not None)

        internal = cells_text(client.get('/api/export/visits.xlsx'))
        assert 'Namit' in internal
        assert 'Secret shelf intel' in internal

        # Owner mode v2 is FAIL-CLOSED: the visits export (rep-level activity
        # detail) is not on the owner allowlist — 403, not a scrubbed copy.
        owner_resp = client.get('/api/export/visits.xlsx?view=owner')
        assert owner_resp.status_code == 403
        assert owner_resp.get_json()['error'] == 'owner view: not permitted'

    def test_top100_export_is_owner_safe_too(self, ingested, client):
        resp = client.get('/api/export/top100.xlsx?view=owner')
        assert resp.status_code == 200
        ws = _load_workbook(resp).active
        text = ' | '.join(str(c.value) for row in ws.iter_rows()
                          for c in row if c.value is not None)
        assert not REP_NAME_RE.search(text)


# ---------------------------------------------------------------------------
# STORE PAGE — contact edits audited + mini reconcile in the full payload
# ---------------------------------------------------------------------------

class TestStorePage:
    @pytest.fixture(scope='class')
    def page_store(self, ingested):
        db = _db()
        sn = db.execute(
            "SELECT store_number FROM territory_stores WHERE tier='routed' "
            "ORDER BY store_number DESC LIMIT 1 OFFSET 3").fetchone()[0]
        store_id = _ensure_store_row(db, sn)
        db.commit()
        db.close()
        return {'sn': sn, 'id': store_id}

    def test_contact_edit_writes_audit_trail(self, page_store, client):
        resp = client.put(f"/api/stores/{page_store['id']}",
                          json={'manager_name': 'Bob Manager',
                                'spirits_ambassador': 'Alice Ambassador',
                                'updated_by': 'Vaneet'})
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['success'] is True
        assert body['audited_fields'] == 2

        db = _db()
        audits = {r['field']: r for r in db.execute(
            "SELECT field, old_value, new_value, changed_by "
            "FROM territory_status_history "
            "WHERE store_number=? AND field LIKE 'contact:%'",
            (page_store['sn'],)).fetchall()}
        db.close()
        assert audits['contact:manager_name']['new_value'] == 'Bob Manager'
        assert audits['contact:manager_name']['old_value'] == ''
        assert audits['contact:manager_name']['changed_by'] == 'Vaneet'
        assert audits['contact:spirits_ambassador']['new_value'] == 'Alice Ambassador'

    def test_unchanged_save_adds_no_audit_noise(self, page_store, client):
        resp = client.put(f"/api/stores/{page_store['id']}",
                          json={'manager_name': 'Bob Manager',
                                'updated_by': 'Vaneet'})
        assert resp.status_code == 200
        assert resp.get_json()['audited_fields'] == 0

    def test_store_full_payload(self, page_store, client):
        body = client.get(f"/api/crm/store/{page_store['sn']}/full").get_json()
        assert body['store']['manager_name'] == 'Bob Manager'

        lu = body['contacts_last_update']  # "last updated by X on date"
        assert lu and lu['changed_by'] == 'Vaneet'
        assert lu['changed_at']

        mini = body['mini_reconcile']  # both SKUs side by side
        assert {m['sku'] for m in mini} == {PHOENIX, DAYAA}
        for m in mini:
            for key in ('sod_on_hand', 'live_qty', 'rep_units', 'flag'):
                assert key in m

        assert isinstance(body['recent_touchpoints'], list)
        t = body['territory']  # top-100 standing on the store page
        assert t is not None
        for key in ('priority_rank', 'owner_status', 'tier', 'route_day'):
            assert key in t

    def test_store_full_owner_view_fails_closed(self, page_store, client):
        # v2: the store page (contacts, notes, rep detail) is not on the
        # owner allowlist — the owner gets 403, nothing to scrub.
        resp = client.get(f"/api/crm/store/{page_store['sn']}/full?view=owner")
        assert resp.status_code == 403
        assert resp.get_json()['error'] == 'owner view: not permitted'
        assert not REP_NAME_RE.search(resp.get_data(as_text=True))


# ---------------------------------------------------------------------------
# OWNER passcode gate — POST /api/owner/check
# ---------------------------------------------------------------------------

class TestOwnerCheck:
    """The /owner gate: passcode is checked server-side against the
    OWNER_PASSCODE env var and never ships in frontend JS. Unset env must
    fail closed (deny everything, even an empty guess)."""

    def test_unset_env_denies_everything(self, client, monkeypatch):
        monkeypatch.delenv('OWNER_PASSCODE', raising=False)
        for guess in ('', 'anything'):
            resp = client.post('/api/owner/check', json={'passcode': guess})
            assert resp.status_code == 200
            assert resp.get_json() == {'ok': False}

    def test_wrong_passcode_denied(self, client, monkeypatch):
        monkeypatch.setenv('OWNER_PASSCODE', 'correct-horse')
        resp = client.post('/api/owner/check', json={'passcode': 'battery-staple'})
        assert resp.get_json() == {'ok': False}

    def test_right_passcode_ok(self, client, monkeypatch):
        monkeypatch.setenv('OWNER_PASSCODE', 'correct-horse')
        resp = client.post('/api/owner/check', json={'passcode': 'correct-horse'})
        assert resp.get_json() == {'ok': True}

    def test_garbage_body_denied_not_500(self, client, monkeypatch):
        monkeypatch.setenv('OWNER_PASSCODE', 'correct-horse')
        resp = client.post('/api/owner/check', data='not json',
                           content_type='application/json')
        assert resp.status_code == 200
        assert resp.get_json() == {'ok': False}


class TestOosRiskOwnerScrub:
    """/api/crm/oos-risk rides on the owner dashboard — rep names in its
    rows must never reach the owner view."""

    def test_owner_view_has_no_rep_names(self, ingested, client):
        resp = client.get('/api/crm/oos-risk?view=owner')
        assert resp.status_code == 200
        assert not REP_NAME_RE.search(resp.get_data(as_text=True))
