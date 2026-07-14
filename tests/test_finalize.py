"""Finalization tests (2026-07-14 punch list) — FIX A + FIX D.

FIX A — owner mode v2:
  - FAIL-CLOSED allowlist: any /api path not on the owner allowlist returns
    403 to X-View: owner (and ?view=owner) — present AND future leaks die.
  - GTA region labels: rep identities become GTA CENTRAL / WEST / NORTH /
    EAST (unknown reps collapse to 'GTA'), including word-boundary
    occurrences inside free-text values.
  - Roster hygiene: the stale legacy reps row 'Ikshit Sharma' never shows
    in roster/filter lists but is NEVER deleted from the DB.

FIX D — data forever:
  - The export/backup surfaces cover every new table (territory_stores,
    territory_status_history, lcbo_live_batches, lcbo_live_snapshots,
    live_listing_events, rep_listing_observations).
  - Retention guard: no DELETE/TRUNCATE against the protected append-only
    tables anywhere in app.py; the SOD snapshot rollback archives instead
    of deleting; import mode=replace refuses to truncate protected tables.
  - DEPLOY_RUNBOOK.md documents the ALERT_EMAIL_TO requirement.

Run with: python3 -m pytest tests/test_finalize.py -v
"""
import io
import json
import os
import re
import sqlite3
import sys
import zipfile

import pytest

# Force SQLite for tests so we never touch production Postgres.
os.environ.pop('DATABASE_URL', None)
os.environ.pop('ADMIN_TOKEN', None)
os.environ.pop('API_KEY', None)
os.environ.pop('SOD_CRON_TOKEN', None)

TEST_DB_DIR = '/tmp/drippcan_finalize_test'
os.makedirs(TEST_DB_DIR, exist_ok=True)
TEST_DB = os.path.join(TEST_DB_DIR, 'drippcan.db')

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

PHOENIX = '0014318'
DAYAA = '0044451'
APP_PY = os.path.join(os.path.dirname(__file__), '..', 'app.py')
RUNBOOK = os.path.join(os.path.dirname(__file__), '..', 'DEPLOY_RUNBOOK.md')

# Word-boundary scan over every roster identity incl. the stale legacy row.
ROSTER_NAME_RE = re.compile(r'\b(Ikshit Sharma|Ikshit|Vaneet|Ed|Namit)\b')

NEW_TABLES = (
    'territory_stores', 'territory_status_history', 'lcbo_live_batches',
    'lcbo_live_snapshots', 'live_listing_events', 'rep_listing_observations',
)


@pytest.fixture(scope='module')
def app_module():
    """Import app.py fresh against an isolated SQLite file (same pattern as
    test_dripp.py — DB_DIR re-asserted because collection order matters)."""
    prev_db_dir = os.environ.get('DB_DIR')
    os.environ['DB_DIR'] = TEST_DB_DIR
    if os.path.exists(TEST_DB):
        os.remove(TEST_DB)
    for mod in list(sys.modules):
        if mod == 'app' or mod.startswith('app.'):
            del sys.modules[mod]
    import importlib.util
    spec = importlib.util.spec_from_file_location('app', APP_PY)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    yield m
    if prev_db_dir is None:
        os.environ.pop('DB_DIR', None)
    else:
        os.environ['DB_DIR'] = prev_db_dir


@pytest.fixture
def client(app_module):
    app_module.app.config['TESTING'] = True
    app_module._rate_buckets.clear()
    with app_module.app.test_client() as c:
        yield c


def _db():
    conn = sqlite3.connect(TEST_DB)
    conn.row_factory = sqlite3.Row
    return conn


@pytest.fixture(scope='module')
def ingested(app_module):
    app_module.app.config['TESTING'] = True
    with app_module.app.test_client() as c:
        resp = c.post('/api/territory/ingest')
        assert resp.status_code == 200, resp.get_data(as_text=True)
        return resp.get_json()


@pytest.fixture(scope='module')
def seeded(app_module, ingested):
    """Plant data on every leak path: an Ed touchpoint with a name-dropping
    note, an Ed shelf observation, a stale 'Ikshit Sharma' store assignment
    and activity actor, and SOD presence for the reconcile board."""
    db = _db()
    stores = [r[0] for r in db.execute(
        "SELECT store_number FROM territory_stores WHERE active=1 "
        "ORDER BY store_number LIMIT 3").fetchall()]
    s_ed, s_stale, s_extra = stores

    def store_id_for(sn):
        row = db.execute(
            "SELECT id FROM stores WHERE store_number=?", (sn,)).fetchone()
        if row is None:
            db.execute("INSERT INTO stores (store_number, account) VALUES (?,?)",
                       (sn, f'LCBO #{sn}'))
            row = db.execute(
                "SELECT id FROM stores WHERE store_number=?", (sn,)).fetchone()
        return row[0]

    ed_id = db.execute("SELECT id FROM reps WHERE name='Ed'").fetchone()[0]
    stale_id = db.execute(
        "SELECT id FROM reps WHERE name='Ikshit Sharma'").fetchone()[0]

    # Ed's touchpoint with a note that names two reps (word-boundary target)
    db.execute(
        "INSERT INTO activities (store_id, rep_id, rep, activity_type, notes, created_at) "
        "VALUES (?,?,?,?,?,?)",
        (store_id_for(s_ed), ed_id, 'Ed', 'visit',
         'Ed dropped samples; Namit to follow up next week',
         '2026-07-13 10:00:00'))
    # Stale-actor activity (must never surface as a fifth rep)
    db.execute(
        "INSERT INTO activities (store_id, rep_id, rep, activity_type, created_at) "
        "VALUES (?,?,?,?,?)",
        (store_id_for(s_stale), stale_id, 'Ikshit Sharma', 'visit',
         '2026-07-13 11:00:00'))
    # Stale store assignment (reps-with-stores filter list)
    db.execute("UPDATE stores SET rep='Ikshit Sharma' WHERE store_number=?",
               (s_stale,))
    db.execute("UPDATE stores SET rep='Ed' WHERE store_number=?", (s_ed,))
    # Ed's shelf observation with a note (rides /api/reconcile)
    db.execute(
        "INSERT INTO rep_listing_observations (sku, store_number, rep, on_shelf, units, notes) "
        "VALUES (?,?,?,?,?,?)",
        (PHOENIX, s_ed, 'Ed', 1, 2, 'Ed says the end cap is empty'))
    # SOD presence so reconcile/top100 boards have real rows
    for sn, oh in ((s_ed, 10), (s_stale, 4), (s_extra, 6)):
        db.execute(
            "INSERT OR IGNORE INTO sod_inventory "
            "(sku, store_number, snapshot_date, status, on_hand, product_name) "
            "VALUES (?,?,?,?,?,?)",
            (PHOENIX, sn, '2026-07-13', 'L', oh, 'PHOENIX ULTRA SMOOTH VODKA'))
    db.commit()
    db.close()
    return {'ed_store': s_ed, 'stale_store': s_stale}


OWNER_HDR = {'X-View': 'owner'}

# Every GET surface the owner IS allowed to reach (nocache so the seeded
# data is always visible, never a stale pre-seed copy).
OWNER_ALLOWED_GETS = [
    '/api/territory?nocache=1',
    '/api/top100?nocache=1',
    '/api/top100/funnel?nocache=1',
    '/api/changes?days=60&nocache=1',
    '/api/conversion?days=60&nocache=1',
    f'/api/reconcile?days=7&sku={PHOENIX}&nocache=1',
    '/api/crm/oos-risk?nocache=1',
    '/api/crm/dashboard?nocache=1',
    f'/api/crm/sku-trend/{PHOENIX}?nocache=1',
    '/api/export/top100.xlsx',
    '/api/export/territory.xlsx',
    '/api/export/changes.xlsx?days=60',
    '/api/export/reconcile.xlsx?days=7',
]

# Internal-only surfaces that must 403 in owner view — the spec's three plus
# a sweep of the other known leak paths (roster, manager, admin, sod, logs).
OWNER_FORBIDDEN_GETS = [
    '/api/crm/rep-performance',
    '/api/crm/activities',
    '/api/admin/export',
    '/api/crm/daily-log',
    '/api/crm/manager-dashboard',
    '/api/reps',
    '/api/crm/reps-with-stores',
    '/api/dashboard',
    '/api/export/visits.xlsx',
    '/api/admin/export/everything',
    '/api/territory/discovery',
]


class TestOwnerFailClosed:
    """FIX A: the owner view is an allowlist, not a blocklist."""

    @pytest.mark.parametrize('path', OWNER_FORBIDDEN_GETS)
    def test_owner_403_on_internal_endpoints(self, ingested, seeded, path, client):
        resp = client.get(path, headers=OWNER_HDR)
        assert resp.status_code == 403, f'{path} must fail closed for the owner'
        assert resp.get_json() == {'error': 'owner view: not permitted'}

    def test_query_param_view_owner_fails_closed_too(self, ingested, client):
        resp = client.get('/api/crm/rep-performance?view=owner')
        assert resp.status_code == 403
        assert resp.get_json() == {'error': 'owner view: not permitted'}

    def test_owner_writes_outside_the_two_allowed_are_403(self, ingested, client):
        # Territory ingest + discovery-add are internal-only writes
        assert client.post('/api/territory/ingest',
                           headers=OWNER_HDR).status_code == 403
        assert client.post('/api/territory/discovery/add',
                           json={'store_number': 1},
                           headers=OWNER_HDR).status_code == 403

    @pytest.mark.parametrize('path', OWNER_ALLOWED_GETS)
    def test_owner_200_on_each_allowlisted_path(self, ingested, seeded, path, client):
        resp = client.get(path, headers=OWNER_HDR)
        assert resp.status_code == 200, f'{path} must stay reachable for the owner'

    def test_owner_check_and_the_two_owner_writes_allowed(self, ingested, client):
        resp = client.post('/api/owner/check', json={'passcode': 'x'},
                           headers=OWNER_HDR)
        assert resp.status_code == 200  # gate answers (ok:false without env)

        db = _db()
        sn = db.execute("SELECT store_number FROM territory_stores "
                        "WHERE active=1 LIMIT 1").fetchone()[0]
        db.close()
        assert client.post('/api/top100/priority',
                           json={'store_number': sn, 'rank': 3},
                           headers=OWNER_HDR).status_code == 200
        assert client.post('/api/top100/status',
                           json={'store_number': sn,
                                 'owner_status': 'listing_received'},
                           headers=OWNER_HDR).status_code == 200

    def test_healthz_is_never_blocked(self, ingested, client):
        resp = client.get('/healthz', headers=OWNER_HDR)
        assert resp.status_code != 403  # 200 or 503-stale, never fail-closed

    def test_internal_view_is_unaffected(self, ingested, seeded, client):
        for path in ('/api/crm/rep-performance', '/api/reps', '/api/dashboard'):
            assert client.get(path).status_code == 200, path


class TestOwnerRegionLabels:
    """FIX A: rep identities become GTA region labels, never names."""

    def test_sanitize_maps_names_to_region_labels(self, app_module, ingested):
        with app_module.app.test_request_context('/?view=owner'):
            out = app_module._owner_sanitize({
                'rep': 'Ed',
                'changed_by': 'Ikshit Sharma',
                'observed_by': 'Somebody Unknown',
                'summary': 'Ed dropped samples; Namit to follow up',
            })
        assert out['rep'] == 'GTA NORTH'
        assert out['changed_by'] == 'GTA CENTRAL'   # stale alias -> same region
        assert out['observed_by'] == 'GTA'          # unknown rep -> plain GTA
        # Word-boundary in-string replacement uses the region label too
        assert out['summary'] == 'GTA NORTH dropped samples; GTA EAST to follow up'

    def test_sanitize_word_boundary_and_neutral_actors(self, app_module, ingested):
        with app_module.app.test_request_context('/?view=owner'):
            out = app_module._owner_sanitize({
                'changed_by': 'seed-default',
                'added_by': 'owner',
                'headline': 'Education aisle refreshed in Vaughan',
            })
        assert out['changed_by'] == 'seed-default'  # neutral actors untouched
        assert out['added_by'] == 'owner'
        # 'Ed' inside 'Education' must NOT be replaced (word boundary)
        assert out['headline'] == 'Education aisle refreshed in Vaughan'

    def test_owner_reconcile_shows_region_not_name(self, ingested, seeded, client):
        resp = client.get(
            f'/api/reconcile?days=7&sku={PHOENIX}&view=owner&nocache=1')
        assert resp.status_code == 200
        text = resp.get_data(as_text=True)
        assert 'GTA NORTH' in text, "Ed's observation must show as GTA NORTH"
        assert not ROSTER_NAME_RE.search(text)

    @pytest.mark.parametrize('path', OWNER_ALLOWED_GETS)
    def test_no_roster_name_in_any_allowlisted_owner_payload(
            self, ingested, seeded, path, client):
        resp = client.get(path, headers=OWNER_HDR)
        assert resp.status_code == 200
        if path.startswith('/api/export/'):
            from openpyxl import load_workbook
            ws = load_workbook(io.BytesIO(resp.data)).active
            text = ' | '.join(str(c.value) for row in ws.iter_rows()
                              for c in row if c.value is not None)
        else:
            text = resp.get_data(as_text=True)
        hit = ROSTER_NAME_RE.search(text)
        assert not hit, f'roster name {hit and hit.group(0)!r} leaked in {path}'


class TestRosterHygiene:
    """FIX A: the stale 'Ikshit Sharma' row is filtered, never deleted."""

    def test_stale_row_stays_in_db_but_out_of_api_reps(self, ingested, seeded, client):
        body = client.get('/api/reps').get_json()
        names = {r['name'] for r in body}
        assert names == {'Ikshit', 'Vaneet', 'Ed', 'Namit'}
        assert 'Ikshit Sharma' not in names
        # NOTHING deleted: the legacy row is still in the table
        db = _db()
        n = db.execute(
            "SELECT COUNT(*) FROM reps WHERE name='Ikshit Sharma'").fetchone()[0]
        db.close()
        assert n == 1

    def test_reps_with_stores_excludes_stale(self, ingested, seeded, client):
        body = client.get('/api/crm/reps-with-stores').get_json()
        reps = {r['rep'] for r in body}
        assert 'Ikshit Sharma' not in reps
        assert 'Ed' in reps  # real roster assignments still counted

    def test_rep_performance_is_roster_only(self, ingested, seeded, client):
        body = client.get('/api/crm/rep-performance?days=30&nocache=1').get_json()
        reps = [e['rep'] for e in body['reps']]
        assert sorted(reps) == sorted(['Ikshit', 'Vaneet', 'Ed', 'Namit'])
        assert 'Ikshit Sharma' not in reps

    def test_dashboard_by_rep_excludes_stale(self, ingested, seeded, client):
        body = client.get('/api/dashboard').get_json()
        names = set(body['by_rep'].keys())
        assert 'Ikshit Sharma' not in names
        assert names == {'Ikshit', 'Vaneet', 'Ed', 'Namit'}


class TestBackupCompleteness:
    """FIX D: every new table rides every backup/export surface."""

    def test_export_tables_include_all_new_tables(self, app_module):
        names = {t for t, _ in app_module._EXPORT_TABLES}
        missing = set(NEW_TABLES) - names
        assert not missing, f'_EXPORT_TABLES is missing: {sorted(missing)}'

    def test_admin_export_core_includes_new_tables(self, ingested, seeded, client):
        body = client.get('/api/admin/export?include=core').get_json()
        for t in NEW_TABLES:
            assert t in body['tables'], f'{t} missing from /api/admin/export'
            td = body['tables'][t]
            assert 'error' not in td, f'{t} export errored: {td}'
            assert 'row_count' in td
        # The territory book actually has rows in it (not a vacuous pass)
        assert body['tables']['territory_stores']['row_count'] >= 189

    def test_export_everything_zip_includes_new_tables(self, ingested, seeded, client):
        resp = client.get('/api/admin/export/everything',
                          headers={'Origin': 'http://localhost:3002'})
        assert resp.status_code == 200
        zf = zipfile.ZipFile(io.BytesIO(resp.data))
        names = set(zf.namelist())
        for t in NEW_TABLES:
            assert f'{t}.csv' in names, f'{t}.csv missing from everything.zip'
        manifest = json.loads(zf.read('manifest.json'))
        for t in NEW_TABLES:
            assert t in manifest['tables']

    def test_essential_email_backup_covers_new_tables(self, app_module, ingested, seeded):
        with app_module.app.test_request_context('/'):
            payload = app_module._build_essential_backup()
        for t in NEW_TABLES:
            assert t in payload['tables'], f'{t} missing from the daily email backup'
            assert 'error' not in payload['tables'][t]
        # The giant snapshot tables stay out of the email (size), as designed
        assert 'sod_inventory' not in payload['tables']
        assert 'inventory_history' not in payload['tables']


class TestRetentionGuard:
    """FIX D: protected tables can never be hard-deleted or truncated."""

    PROTECTED = ('sod_inventory', 'lcbo_live_snapshots', 'activities',
                 'territory_status_history')

    def test_no_hard_deletes_on_protected_tables_in_source(self):
        src = open(APP_PY).read()
        pattern = re.compile(
            r"(?:DELETE\s+FROM|TRUNCATE(?:\s+TABLE)?)\s+(?:%s)\b"
            % '|'.join(self.PROTECTED), re.IGNORECASE)
        hits = pattern.findall(src)
        assert hits == [], (
            f'hard DELETE/TRUNCATE against protected tables found: {hits} — '
            'use _archive_then_remove / soft-delete instead')

    def test_protected_set_matches_the_spec(self, app_module):
        assert set(self.PROTECTED) <= set(app_module._RETENTION_PROTECTED_TABLES)

    def test_rollback_snapshot_archives_instead_of_deleting(self, ingested, client):
        db = _db()
        for sn in (9801, 9802, 9803):
            db.execute(
                "INSERT OR IGNORE INTO sod_inventory "
                "(sku, store_number, snapshot_date, status, on_hand, product_name) "
                "VALUES (?,?,?,?,?,?)",
                (DAYAA, sn, '2020-01-01', 'L', 5, 'DAYAA ARAK'))
        db.execute(
            "INSERT OR IGNORE INTO sod_inventory "
            "(sku, store_number, snapshot_date, status, on_hand, product_name) "
            "VALUES (?,?,?,?,?,?)",
            (DAYAA, 9801, '2020-01-02', 'L', 5, 'DAYAA ARAK'))
        db.commit()
        db.close()

        # Refuses without confirm
        resp = client.post('/api/admin/sod/rollback-snapshot',
                           json={'snapshot_date': '2020-01-01'},
                           headers={'Origin': 'http://localhost:3002'})
        assert resp.status_code == 400

        resp = client.post('/api/admin/sod/rollback-snapshot',
                           json={'snapshot_date': '2020-01-01', 'confirm': True},
                           headers={'Origin': 'http://localhost:3002'})
        assert resp.status_code == 200, resp.get_data(as_text=True)
        body = resp.get_json()
        assert body['archived_rows'] == 3

        db = _db()
        live = db.execute(
            "SELECT COUNT(*) FROM sod_inventory WHERE snapshot_date='2020-01-01'"
        ).fetchone()[0]
        other = db.execute(
            "SELECT COUNT(*) FROM sod_inventory WHERE snapshot_date='2020-01-02'"
        ).fetchone()[0]
        archived = db.execute(
            "SELECT COUNT(*) FROM sod_inventory_archive "
            "WHERE snapshot_date='2020-01-01'").fetchone()[0]
        db.close()
        assert live == 0        # gone from the live board
        assert other == 1       # only the requested date moved
        assert archived == 3    # ...and every row lives on in the archive

    def test_import_replace_never_truncates_protected_tables(self, ingested, seeded, client):
        db = _db()
        before = db.execute("SELECT COUNT(*) FROM activities").fetchone()[0]
        store_id = db.execute("SELECT id FROM stores LIMIT 1").fetchone()[0]
        rep_id = db.execute("SELECT id FROM reps WHERE name='Ed'").fetchone()[0]
        db.close()
        assert before >= 1  # seeded fixture planted activity rows

        payload = {'tables': {'activities': {
            'columns': ['id', 'store_id', 'rep_id', 'activity_type'],
            'rows': [{'id': 987654, 'store_id': store_id, 'rep_id': rep_id,
                      'activity_type': 'visit'}],
        }}}
        resp = client.post('/api/admin/import?mode=replace&confirm=YES',
                           json=payload)
        assert resp.status_code == 200
        table_result = resp.get_json()['tables']['activities']
        assert table_result['mode'] == 'merge'  # retention guard downgraded it

        db = _db()
        after = db.execute("SELECT COUNT(*) FROM activities").fetchone()[0]
        db.close()
        assert after == before + 1  # nothing truncated, new row landed

    def test_runbook_documents_alert_email_to(self):
        text = open(RUNBOOK).read()
        assert 'ALERT_EMAIL_TO' in text
        assert '02:00' in text          # the daily backup time is documented
        assert 'PITR' in text           # Neon free-tier window is documented
